import random as r

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('NBAplay.xlsx')

ws= wb.active

class NBA:
    #Class Attributes (Apply to all instances of the class)
    ball_size = 29.5
    ball_type = 'Wilson'
    name = 'National Basketball Association'

    def __init__(self, player, team, pos, no, college):
        #Instance Attributes (Different for every instance, but every instance has to have these)
        #Used for guessing the players, except player is the player you are trying to guess
        self.player = player
        self.team = team
        self.pos = pos
        self.no = no
        self.college = college

#Function that creates an instance (object) of NBA to use in the game
def PlayerPicker(row):
    players = [] #Empty list (this is where the attributes of the instance will go from the excel file)
    name = ws[row] #Pull specific row (player) from excel sht
    for i in name:
        players.append(i.value) #Append each row value to players list
    global SecretPlay
    SecretPlay = NBA(players[1], players[3], players[2], players[0], players[4]) #Object that spells out each attribute of the class

num =1
playChoiceInt = [] #Pick which player by choosing a row 

while len(playChoiceInt) < 474: #Add numbers to list for players
    playChoiceInt.append(num)
    num += 1 
    
#Converting interger list to string for pyopenxl to read
playChoice = [str(x) for x in playChoiceInt] 

PlayerPicker(r.choice(playChoice)) #Random module to make it random, insert it into the function

g=1 #Used to advance the game to the next screen
while input() != SecretPlay.player: #If input does not equal the players name, skips to the next one
    if g == 1:
        print('He plays for the '+ SecretPlay.team)
        g=2
    elif g ==2:
        print('His position is '+ SecretPlay.pos) 
        g=3
    elif g ==3 :
        print ('His number is '+ str(SecretPlay.no))
        g=4
    elif g == 4:
        if SecretPlay.college == 'None':
            print('He did not go to college')
            g=5
        else:
            print('He went to ' + SecretPlay.college)
            g=5
    elif g==5: #If you can't get it by here, you end
        print('Nope! Better Luck Next Time\nThe Player was '+ SecretPlay.player ) 
        break
else: #They got the right answer in time
    print('Correct!')
    


        

  
